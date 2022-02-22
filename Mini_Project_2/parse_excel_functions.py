import re
import openpyxl
import logging
import validation_functions as validate
import file_functions as ff


class ProgramError(Exception):

    """
    custom error to handle if anything fails during parsing of excel file
    error is called after known handled exceptions so the program does not crash
    errors are logged and then printed to the user for information on what happened
    """
    pass


# main function to parse target excel data from workbook
def parse_excel_data(file_name_to_parse: str):

    """
    Todo: write detailed discription of what this function does
    """

    # try block if anything goes wrong with parsing of file logs and displays to User
    # this keeps from program reporting wrong details at the end or crashing
    try:

        # splits file name on "_" and "." into list to validate
        file_name_list = re.split('_|\\.', file_name_to_parse)
        month_from_filename = file_name_list[3]
        year_from_filename = file_name_list[4]

        # loads excel workbook
        wb_obj = openpyxl.load_workbook(
            "/Users/stephenfreed/Projects/SmoothStack/Mini_Project_2/excel_files/{}".format(file_name_to_parse)
            )

        # runs function to validate proper sheet present in workbook / returns tuple
        sheet_present_tuple = validate.check_for_sheet(wb_obj)

        # if return tuple has not validated sheet raises ProgramError
        if sheet_present_tuple[0] is False:
            raise ProgramError

        # if valid sets sheet index in list and proper sheet object to parse from
        sheet_name = sheet_present_tuple[1]
        sheet_obj = wb_obj[sheet_name]

        # finds column number of proper month and year(if present in cell) to parse from
        # 2 valid cell scenarios: full month spelled out or datetime
        column_to_parse = None
        m_column = sheet_obj.max_column

        # loop through top row looking for proper date (2 valid choices)
        for column_number in range(1, m_column + 1):

            try:
                # gets header cell info and turns it into a string
                header_cell_obj = sheet_obj.cell(row=1, column=column_number)
                cell_string = str(header_cell_obj.value).strip()

                # skips blank cells
                if header_cell_obj.value is None:
                    continue

                # calls both types of validation functions to see if either is True
                datetime_result = validate.validate_datetime(cell_string, month_from_filename, year_from_filename)
                month_string_result = validate.validate_month_string(cell_string, month_from_filename)

                # if either are true sets column number to parse from and breaks loop
                if datetime_result is True or month_string_result is True:
                    column_to_parse = column_number
                    break

            except: # noqa
                logging.error("While Finding Matching Column Name In parse_excel_functions.py")
                print("\n(ERROR) While Finding Matching Month and Year(if present) In parse_excel_functions.py")
                raise ProgramError

        # checks if column_to_parse was assigned
        if column_to_parse is None:
            logging.error("No Valid Matching Month and Year(if present) Cell Was Found In parse_excel_functions.py")
            print("\n(ERROR) No Valid Matching Month and Year(if present) Cell Was Found In parse_excel_functions.py")
            raise ProgramError

        # find rows to parse by building list of tuples (row_name, row_number)
        rows_to_parse_list = []
        m_row = sheet_obj.max_row

        # loop through first column looking for correct rows to parse (3 choices)
        for row_number in range(1, m_row + 1):

            try:
                # gets cell info and turns it into a string
                cell_obj = sheet_obj.cell(row=row_number, column=1)  # type: ignore
                cell_string = str(cell_obj.value).strip()  # object to string

                # skips blank cells
                if cell_obj.value is None:
                    continue

                # split string into list on space
                split_string_list = re.split(' ', cell_string)

                # check if start of cell is rows we are looking for
                if (
                    split_string_list[0].title() == "Promoters" or
                    split_string_list[0].title() == "Passives" or
                    split_string_list[0].title() == "Dectractors"
                     ):
                    rows_to_parse_list.append((split_string_list[0].title(), row_number))

            except:  # noqa
                logging.error("While Finding Matching Target Rows In parse_excel_functions.py")
                print("\n(ERROR) While Finding Matching Target Rows In parse_excel_functions.py")
                raise ProgramError

        # checks that all target rows are likely found correctly
        if len(rows_to_parse_list) != 3:
            logging.error("rows_to_parse_list Is Not The Correct Length In parse_excel_functions.py")
            print("\n(ERROR) rows_to_parase_list Is of Wrong Length - Likely Missing Target Rows To Parse")
            raise ProgramError

        # builds logging output for successful parsing of data
        successful_log_string = ""
        successful_log_string += "\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n"
        successful_log_string += "Successfully Parsed Excel File: {}\n".format(file_name_to_parse)
        successful_log_string += "Data From Worksheet: {}\n".format(sheet_name)
        successful_log_string += (
                                  "~~~ For the Month of {} in {} ~~~\n"
                                  .format(month_from_filename.capitalize(), year_from_filename)
                                  )

        # loops through target rows to parse list returning data
        for row_name, row_number in rows_to_parse_list:  # unpacks tuple

            try:
                target_row_column_cell_obj = sheet_obj.cell(row=row_number, column=column_to_parse)  # type: ignore
                target_row_column_cell = str(target_row_column_cell_obj.value).strip()
                # check what level
                row_rating = validate.check_row_rating(row_name, target_row_column_cell)
                successful_log_string += "{}:{} {}\n".format(row_name, target_row_column_cell, row_rating)
            except: # noqa
                logging.error("Error While Building successful_log_string In parse_excel_functions.py")
                print("\n(ERROR) Error While Building successful_log_string In parse_excel_functions.py")
                raise ProgramError

        successful_log_string += "~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n"

    except ProgramError:
        logging.error("A Known Exception Was Handled | Application Could Not Finish")
        logging.info(f"File:{file_name_to_parse} Was Moved To error_files/invalid_file Directory")
        ff.move_invalid_file(file_name_to_parse)
        print(
                "The Known Exception Above Was Handled ^^^\n"
                "Application Could Not Finish\n"
                "Please Check Log File...\n"
                "File Was Moved To error_files/invalid_file Directory\n"
              )

    except Exception as e:
        logging.error("The Program Could Not Get Excel Data Successfully")
        logging.info(f"File:{file_name_to_parse} Was Moved To error_files/invalid_file Directory")
        ff.move_invalid_file(file_name_to_parse)
        print(e)
        print("The Program Could Not Get Excel Data Successfully\nTry Another File\nClosed Application...")

    # displays parsed data to user and logs it if there were no exceptions
    else:

        print("\n~ File Parsed Successfully ~")

        # log and print parse info
        logging.info(successful_log_string)
        print(successful_log_string)

        logging.info("File Parsed Successfully")

        # moves file from excel_files to logging/processed_files directory
        # adds entry in files_processed.txt
        ff.move_processed_file(file_name_to_parse)

