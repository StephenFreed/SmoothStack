import re
import time
import datetime
import subprocess
import openpyxl
import logging


# custom error to handle if anything fails during function call
# error is called after known handled exceptions to show user and stop function
class ProgramError(Exception):
    pass


def parse_excel_data(file_name_to_parse: str):

    """
    Parses data based on file name input. Looks for month and year.
    Returns data from that month and year in the excel file.
    """

    # splits file name on "_" and "." into list to validate
    file_name_list = re.split('_|\\.', file_name_to_parse)
    month_from_filename = file_name_list[3]
    year_from_filename = file_name_list[4]

    # loads excel workbook
    wb_obj = openpyxl.load_workbook(
        "/Users/stephenfreed/Projects/SmoothStack/Mini_Project_1/Excel_Files/{}".format(file_name_to_parse)
        )

    # gets active sheet of workbook
    sheet_obj = wb_obj.active

    # try block if anything goes wrong with parsing logs and displays to User
    # this keeps from program reporting wrong details at the end
    try:

        # builds list of valid dates and the rows they were in/skips invalid rows
        parsed_date_list = []
        m_row = sheet_obj.max_row  # type: ignore // gets max row number
        for row_number in range(1, m_row + 1):
            cell_obj = sheet_obj.cell(row=row_number, column=1)  # type: ignore
            cell_string = str(cell_obj.value).strip()  # object to string
            date_format = "%Y-%m-%d %H:%M:%S"
            try:
                # checks if cell is not empty and that date is in correct format
                if cell_obj.value is not None and datetime.datetime.strptime(cell_string, date_format):
                    date_and_row = (cell_string, row_number)
                    parsed_date_list.append(date_and_row)
            # passes if strptime function fails on invalid date format
            except:  # noqa
                pass

        # loops through parsed_date_list and finds row we need to parse
        row_of_info = None  # row to parse info from
        for date, row in parsed_date_list:  # unpacks tuple
            try:
                split_date_list = re.split('-| ', date)  # split date into list on "-" and space
                month_to_number_dict = {
                    "january": "01",
                    "february": "02",
                    "march": "03",
                    "april": "04",
                    "may": "05",
                    "june": "06",
                    "july": "07",
                    "august": "08",
                    "september": "09",
                    "october": "10",
                    "november": "11",
                    "december": "12"
                }
                # checks if year from filename matches year in cell
                # checks if month in file name matches month in cell through month_to_number_dict
                if (split_date_list[0] == year_from_filename and
                   split_date_list[1] == month_to_number_dict[month_from_filename]):
                    row_of_info = row  # row number we need to parse
            except: # noqa
                logging.error("Finding Row With Month And Year Failed. May Not Be In File...")
                print("\nWe Did Not Find The Month and Year To Parse In This Excel Workbook")
                raise ProgramError

        # builds list of tuples with header name and column number/skips invalid cells
        header_and_column_list = []
        m_column = sheet_obj.max_column  # type: ignore
        for column_number in range(1, m_column + 1):
            try:
                header_cell_obj = sheet_obj.cell(row=1, column=column_number)  # type: ignore
                cell_string = str(header_cell_obj.value).strip()
                header_and_column_number = (cell_string, column_number)
                if header_cell_obj.value is not None:
                    header_and_column_list.append(header_and_column_number)
            except: # noqa
                logging.error("Error While Building Header,Column List In parse_excel_functions.py")
                print("\n(ERROR) Error While Building Header,Column List In parse_excel_functions.py")
                raise ProgramError

        # builds logging output for successful parsing of data
        list_of_worksheets = wb_obj.sheetnames  # get workbook list of worksheets
        successful_log_string = ""
        successful_log_string += "\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n"
        successful_log_string += "Successfully Parsed Excel File: {}\n".format(file_name_to_parse)
        successful_log_string += "Data From Worksheet: {}\n".format(list_of_worksheets[0])
        successful_log_string += (
                                  "~~~ For the Month of {} in {} ~~~\n"
                                  .format(month_from_filename.capitalize(), year_from_filename)
                                  )

        # loops through target row concatenating header and row info
        for header, column_number in header_and_column_list:  # unpacks tuple
            try:
                target_row_column_cell_obj = sheet_obj.cell(row=row_of_info, column=column_number)  # type: ignore
                target_row_column_cell = str(target_row_column_cell_obj.value).strip()
                if target_row_column_cell[:2] == "0.":  # check if % in cell
                    target_row_column_cell_float = float(target_row_column_cell)
                    successful_log_string += "{}: {:.2%}\n".format(header, target_row_column_cell_float)
                else:
                    target_row_column_cell_int = int(target_row_column_cell)
                    successful_log_string += "{}: {}\n".format(header, target_row_column_cell_int)
            except Exception: # noqa
                logging.error("Error While Building successful_log_string In parse_excel_functions.py")
                print("\n(ERROR) Error While Building successful_log_string In parse_excel_functions.py")
                raise ProgramError

        successful_log_string += "~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n"

    except ProgramError:
        logging.error("A Known Exception Was Handled | Application Could Not Finish")
        print("The Known Exception Above Was Handled ^^^\nApplication Could Not Finish\nPlease Check Log File...")

    except: # noqa
        logging.error("The Program Could Not Get Excel Data Successfully")
        print("The Program Could Not Get Excel Data Successfully\nTry Another File\nClosed Application...")

    # displays parsed data to user and logs it if there were no exceptions
    else:

        print("\n~ Application Was Successful ~")

        # log and print parse info
        logging.info(successful_log_string)
        print(successful_log_string)

        logging.info("Application Successfully Ran")

        print("~~~~~~~~~~~~~~~~~~~~~~\n ~ Opening Log File ~\n~~~~~~~~~~~~~~~~~~~~~~")
        time.sleep(1)
        subprocess.call(["open", "/Users/stephenfreed/Projects/SmoothStack/Mini_Project_1/logging/log_file.txt"])
