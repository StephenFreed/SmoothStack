import os
import subprocess
import re
import time
import datetime
import openpyxl
import validation_functions as validate
import logging

# log file instead of txt

# structures logging formatting to log_file
logging.basicConfig(
    format="%(asctime)s - [%(levelname)s] - %(message)s",
    level=logging.INFO,
    datefmt="%H:%M:%S %m-%d-%Y",
    filename="/Users/stephenfreed/Projects/SmoothStack/Mini_Project_1/log_file.txt"
)


# pulls in list of file names from target excel file directory
excel_directory = "/Users/stephenfreed/Projects/SmoothStack/Mini_Project_1/Excel_Files/"
excel_file_list = os.listdir(excel_directory)


# builds list of available files from that directory to diplay as choices to user
input_list_of_files = "Choose 0 To Quit Application\n"
for i, file_name in enumerate(excel_file_list):
    input_list_of_files += f"Choose {i+1} for: {file_name}\n"


# if everythong is valid parses data
def parse_excel_data(file_name_to_parse):

    # print(file_name_to_parse)
    file_name_list = re.split('_|\\.', file_name_to_parse) # split file name into list
    month_from_filename = file_name_list[3]
    year_from_filename = file_name_list [4]

    wb_obj = openpyxl.load_workbook("/Users/stephenfreed/Projects/SmoothStack/Mini_Project_1/Excel_Files/{}".format(file_name_to_parse))
    sheet_obj = wb_obj.active

    # build list of first colomn values and find row with matching month and year
    m_row = sheet_obj.max_row
    parsed_date_list = []
    for row_number in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row=row_number, column=1)
        cell_string = str(cell_obj.value)
        date_format = "%Y-%m-%d %H:%M:%S"
        try:
            if cell_obj.value is not None and datetime.datetime.strptime(cell_string, date_format):
                # date_string = str(cell_obj.value)
                date_and_row = (cell_string, row_number)
                parsed_date_list.append(date_and_row)
        except:
            pass

    # print("Parsed Date List: \n {}".format(parsed_date_list))

    # loops through parsed date_list and finds row we need to parse
    row_of_info = None
    for date, row in parsed_date_list:
        try:
            split_date_list = re.split('-| ', date) # split date into list
            # print(split_date_list)
            month_to_number_dict = {"january": "01", "february": "02", "march": "03", "april": "04", "may": "05", "june": "06", "july": "07", "august": "08", "september": "9", "october": "10", "november": "11", "december": "11"}
            # print("{} - {}".format(split_date_list[0], year_from_filename))
            # print("{} - {}".format(split_date_list[1], month_to_number_dict[month_from_filename]))
            if split_date_list[0] == year_from_filename and split_date_list[1] == month_to_number_dict[month_from_filename]:
                row_of_info = row
                # print("The row we need info from is: {}".format(row))
        except:
            # todo something went wrong
            print("something went wrong")

    # return row info
    # get column row header numbers to names
    list_of_worksheets = wb_obj.sheetnames
    # print(list_of_worksheets)

    # get header info
    header_and_column_list = []
    m_column = sheet_obj.max_column
    for column_number in range(1, m_column + 1):
        try:
            cell_obj_2 = sheet_obj.cell(row=1, column=column_number)
            cell_string = str(cell_obj_2.value).strip()
            header_and_column_number = (cell_string, column_number)
            if cell_obj_2.value is not None:
                header_and_column_list.append(header_and_column_number)
        except Exception as e:
            print("somethign went wrong 404")
            print(e)

    print(header_and_column_list)
# Calls Offered: 16,91
# Abandon after 30s : 2.32%
# FCR : 86.50%
# DSAT :  14.20%
# CSAT : 78.30%
    successful_log_string = "\n"
    successful_log_string += "\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n"
    successful_log_string += "This is from worksheet {}\n".format(list_of_worksheets[0])
    successful_log_string += "From Excel file {}\n".format(file_name_to_parse)

    for header, column_number in header_and_column_list:
        try:
            cell_obj_3 = sheet_obj.cell(row=row_of_info , column=column_number)
            successful_log_string += f"{header}: {column_number}\n"
            # print(f"{header}-{column_number}")
        except Exception as e:
            print(e)

    successful_log_string += "~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n"
    logging.info(successful_log_string)
    print(successful_log_string)
        



# asks to select file to parse data from
selection = True
while selection:
    try:
        # asks user which file to parse
        file_selection_number = int(input(
            f"\n-- Please Select A File To Parse --\n{input_list_of_files}\nEnter Number: "))

        # checks to make sure file selection is valid
        # 0 quits application
        if file_selection_number == 0:
            logging.info("Application Was Successfully Closed")
            print("\nExiting Parsing Application...")
            selection = False
        # checks number is in correct range of selections
        elif file_selection_number > len(input_list_of_files) or file_selection_number < 1:
            logging.error("Selected File Number Was Out Of Range Of Choices")
            print("\n(Error!) Select A Valid File Number...")
        # calls is_valid_file_name function / if valid runs application and opens log file
        else:
            is_valid_file = validate.is_valid_file_name(excel_file_list[file_selection_number - 1])
            if is_valid_file != "correct":
                print(is_valid_file)
            else:
                parse_excel_data(excel_file_list[file_selection_number - 1])  # runs app
                logging.info("Application Successfully Ran")
                print(f"\nCongratulations!\n\nYou Selected File: {excel_file_list[file_selection_number - 1]}\n")
                print("~~~~~~~~~~~~~~~~~~~~~~\n ~ Opening Log File ~\n~~~~~~~~~~~~~~~~~~~~~~")
                selection = False
    except Exception as e:
        logging.error("Selected File Number Was Out Of Range Of Choices")
        print("\n(Error!) Select A Valid File Number...")
        # print(e)


# time.sleep(1)
# subprocess.call(["open", "/Users/stephenfreed/Projects/SmoothStack/Mini_Project_1/log_file.txt"])
